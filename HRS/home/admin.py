from django.contrib import admin
from django.http import HttpResponse
import openpyxl
# Register your models here.
from .models import Hotel,Amenities,hotelimages,HotelBooking
from openpyxl.utils import get_column_letter 
admin.site.register(Amenities)
admin.site.register(hotelimages)
# admin.site.register(HotelBooking)
admin.site.register(Hotel)



def export_to_excel(modeladmin, request, queryset):
    """
    Export selected HotelBooking records to an Excel file.
    """
    # Create an Excel workbook and worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Hotel Bookings"

    # Add headers (include total_price, phone_number, room_type)
    headers = ['UID', 'Hotel Name', 'User', 'Start Date', 'End Date', 'Booking Type', 'Total Price', 'Phone Number', 'Room Type']
    sheet.append(headers)

    # Add data rows (include total_price, phone_number, room_type)
    for booking in queryset:
        row = [
            str(booking.uid),  # Convert UUID to string
            getattr(booking.hotel, 'hotel_name', 'N/A'),  # Safely get hotel name
            getattr(booking.user, 'username', 'N/A'),  # Safely get username
            booking.start_date.strftime('%Y-%m-%d'),  # Format date as 'YYYY-MM-DD'
            booking.end_date.strftime('%Y-%m-%d'),  # Format date as 'YYYY-MM-DD'
            booking.booking_type,
            booking.total_price if booking.total_price else 'N/A',  # Show total price or N/A
            booking.phone_number if booking.phone_number else 'N/A',  # Show phone number or N/A
            booking.room_type if booking.room_type else 'N/A',  # Show room type or N/A
        ]
        sheet.append(row)

    # Auto-adjust column width
    for col_idx, column_cells in enumerate(sheet.columns, 1):  # col_idx is 1-based index
        max_length = 0
        column = get_column_letter(col_idx)  # Get the corresponding column letter
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = max_length + 2
        sheet.column_dimensions[column].width = adjusted_width

    # Set up HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=hotel_bookings.xlsx'

    # Save workbook to response
    workbook.save(response)
    return response


class HotelBookingAdmin(admin.ModelAdmin):
    # Add the new fields to the list_display
    list_display = ('uid', 'hotel', 'user', 'start_date', 'end_date', 'booking_type', 'total_price', 'phone_number', 'room_type')  # Updated to show new fields
    actions = [export_to_excel]  # Add export action

admin.site.register(HotelBooking, HotelBookingAdmin)